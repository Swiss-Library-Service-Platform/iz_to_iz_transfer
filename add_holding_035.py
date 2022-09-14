##############################################################
# Add 035 field to holdings when items are left in source IZ #
##############################################################

# To start the script:
# python add_holding_035.py <file_name_processing.csv>
# For example:
# python add_holding_035.py

# Import libraries

from almapiwrapper.inventory import IzBib, NzBib, Holding, Item
from almapiwrapper.configlog import config_log
import sys
import logging
import openpyxl
import pandas as pd
from lxml import etree

# Config logs
config_log(sys.argv[1].replace('\\', '/').split('/')[-1].split('.')[0] + '_add_035')

if len(sys.argv) != 2:
    logging.critical('Argument missing or not correct')

# Get IDs of the processed holdings
process_file_path = 'data/' + sys.argv[1].replace('\\', '/').split('/')[-1].split('.')[0] + '_processing.csv'

# Get configuration
wb = openpyxl.load_workbook(sys.argv[1])
wb.active = wb['General']
sheet = wb.active
iz_s = sheet.cell(row=3, column=2).value
iz_d = sheet.cell(row=4, column=2).value
env = {'Production': 'P',
       'Sandbox': 'S'}.get(sheet.cell(row=5, column=2).value, 'P')

# Load holdings data
df = pd.read_csv(process_file_path)

# Display introduction
print(f'''

##############################################################
# Add 035 field to holdings when items are left in source IZ #
##############################################################

Configuration
=============

Environment: {env}
Source IZ: {iz_s}
Destination IZ: {iz_d}
Nb holdings: {len(df)}

Start job
=========

''')

# Load a holdings
for row in list(df.iterrows()):
    holding = Holding(mms_id=row[1]['MMS_id_s'], holding_id=row[1]['Holding_id_s'], zone=iz_s, env=env)
    new_035_field = f'(IZ-{iz_d}){row[1]["Holding_id_d"]}'

    # Check if the field already existst, if yes, skip the holding
    if len([field for field in holding.data.findall('.//datafield[@tag="035"]/subfield[@code="a"]')
            if field.text == new_035_field]) > 0:
        logging.warning(f'{repr(holding)}: 035 field with "{new_035_field}" text already existing')
        continue

    # Get items
    items = holding.get_items()

    # Check if all item have "OLD_" predfix on barcode, if yes, skip the holding
    if len([item for item in items if item.barcode.startswith('OLD_') is False]) == 0:
        continue

    # Update the holdings
    holding.data.find('.//record').append(etree.XML(
        f'<datafield tag="035" ind1=" " ind2=" "><subfield code="a">{new_035_field}</subfield></datafield>'))
    holding._data.sort_fields()
    logging.info(f'{repr(holding)}: 035a field added "{new_035_field}"')
    holding.update()
