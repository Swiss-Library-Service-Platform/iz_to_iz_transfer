##############################################################
# Build a csv list with items information of the destination #
##############################################################

# This script use stored records to build a csv list with information about the item. This csv file
# can be used to transfer in the Speibi LVS System.
# Fields are: "type / creator / imprint", "title", "barcode",
# "item call number", "material type", "description", "call number"

# The information about modification should be given in a Excel file
# This file should be compliant with a given format

# To start the script:
# python build_csv_item_list.py <dataForm.xlsx>

# Import libraries

from almapiwrapper.configlog import config_log
from almapiwrapper.inventory import Item
from almapiwrapper.record import XmlData
import sys
import logging
import openpyxl
import pandas as pd
import os
from datetime import date
from lxml import etree

# Config logs
config_log(sys.argv[1].replace('\\', '/').split('/')[-1].split('.')[0] + '_add_035')

if len(sys.argv) != 2:
    logging.critical('Argument missing or not correct')

# Get IDs of the processed holdings
process_file_path = 'data/' + sys.argv[1].replace('\\', '/').split('/')[-1].split('.')[0] + '_processing.csv'

# Get configuration
wb = openpyxl.load_workbook(sys.argv[1])
wb.active = wb['General']
sheet = wb.active
iz_s = sheet.cell(row=3, column=2).value
iz_d = sheet.cell(row=4, column=2).value
env = {'Production': 'P',
       'Sandbox': 'S'}.get(sheet.cell(row=5, column=2).value, 'P')

# Load holdings data
cols = {'Barcode': str,
        'NZ_mms_id': str,
        'MMS_id_s': str,
        'Holding_id_s': str,
        'Item_id_s': str,
        'MMS_id_d': str,
        'Holding_id_d': str,
        'Item_id_d': str,
        'Process': str,
        'Copied': bool,
        'Error': str}
df = pd.read_csv(process_file_path, dtype=cols)

# Display introduction
print(f'''

##############################################################
# Build a csv list with items information of the destination #
##############################################################

Configuration
=============

Environment: {env}
Source IZ: {iz_s}
Destination IZ: {iz_d}
Nb items: {len(df)}
Nb items copied: {len(df.loc[df["Copied"]])}

Only copied items will be in the list!

Start job
=========

''')
d_data = pd.DataFrame(columns=["type / creator / imprint",
                               "title",
                               "barcode",
                               "item call number",
                               "material type",
                               "description",
                               "call number",
                               "MMS ID",
                               "Holdings ID",
                               "Item ID",
                               "Library",
                               "Location"])

# Filter data with only copied records
df = df.loc[df["Copied"]]

# Build table
for i, row in list(df.iterrows()):
    mms_id = row['MMS_id_d']
    holding_id = row['Holding_id_d']
    item_id = row['Item_id_d']
    records = os.listdir('records')
    item = None
    if f'{iz_d}_{mms_id}' in records:
        items = [item for item in os.listdir(f'records/{iz_d}_{mms_id}') if f'item_{holding_id}_{item_id}' in item]

        if len(items) > 0:
            items.sort()
            item_filename = items[-1]
            data_path = f'records/{iz_d}_{mms_id}/{item_filename}'
            data = XmlData(filepath=data_path)
            item = Item(mms_id, holding_id, item_id, iz_d, env, data=data)
    if item is None:
        item = Item(mms_id, holding_id, item_id, iz_d, env)
        logging.warning(f'{repr(item)}: no local data found')
        _ = item.data # Load data

    if item.error is False:
        logging.info(f'{repr(item)}: item data available and stored')
        fields = ['bib_data/author',
                  'bib_data/title',
                  'item_data/barcode',
                  'item_data/alternative_call_number',
                  'item_data/physical_material_type',
                  'item_data/description',
                  'holding_data/call_number',
                  'bib_data/mms_id',
                  'holding_data/holding_id',
                  'item_data/pid',
                  'item_data/library',
                  'item_data/location']

        row = [item.data.find(field).text if item.data.find(field) is not None else ''
               for field in fields]
        d_data.loc[len(d_data)] = row
for library in d_data['Library'].unique():
    for location in d_data.loc[d_data['Library'] == library]['Location'].unique():
        file_name = f'data/{library}_{location}_NEW_{date.today().strftime("%Y%m%d")}.csv'
        d_data.loc[(d_data['Library'] == library)&(d_data['Location'] == location)].to_csv(file_name, sep=';', index=False)
        logging.info(f'File "{file_name}" saved')
