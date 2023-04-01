#####################
# Transfer IZ to IZ #
#####################

# This script transfers item from IZ source to IZ destination.
# The information about the transfer should be given in a Excel file
# This file should be compliant with a given format

# To start the script:
# python transfer_iz_to_iz.py <dataForm.xlsx>

# Import libraries
from almapiwrapper.inventory import IzBib, Holding, Item
from almapiwrapper.configlog import config_log
import pandas as pd
from copy import deepcopy
import sys
import os
import logging
import openpyxl
import time

# Config logs
config_log(sys.argv[1].replace('\\', '/').split('/')[-1].split('.')[0])

FORCE_COPY = False
FORCE_UPDATE = False

# Item fields blocking copy like provenance are removed
if len(sys.argv) == 3 and sys.argv[2] == '--force':
    FORCE_COPY = True
    FORCE_UPDATE = True

# All fields of the item are preserved
elif len(sys.argv) == 2:
    FORCE_COPY = False

# Bad argument, program stop
else:
    logging.critical('Argument missing or not correct')
    exit()

# File path to the backup of the processed records
process_file_path = 'data/' + sys.argv[1].replace('\\', '/').split('/')[-1].split('.')[0] + '_processing.csv'

# Get configuration
wb = openpyxl.load_workbook(sys.argv[1])
wb.active = wb['General']
sheet = wb.active
iz_s = sheet.cell(row=3, column=2).value
iz_d = sheet.cell(row=4, column=2).value
env = {'Production': 'P',
       'Sandbox': 'S'}.get(sheet.cell(row=5, column=2).value, 'P')

# Load barcodes
barcodes = pd.read_excel(sys.argv[1], sheet_name=1, dtype=str)['Barcode'].dropna().str.strip("'")
logging.info(f'{len(barcodes)} barcodes loaded from "{sys.argv[1]}" file.')

# Load locations
locations_table = pd.read_excel(sys.argv[1], sheet_name=2, dtype=str)

# Load item policies
item_policies_table = pd.read_excel(sys.argv[1], sheet_name=3, dtype=str)

# Check if processing file exists
if os.path.exists(process_file_path) is True:
    df = pd.read_csv(process_file_path, dtype=str)
    df = df.replace('False', False)
    df = df.replace('True', True)
    df = df.replace('NaN', None)

else:
    df = pd.DataFrame(columns=['Barcode',
                               'NZ_mms_id',
                               'MMS_id_s',
                               'Holding_id_s',
                               'Item_id_s',
                               'MMS_id_d',
                               'Holding_id_d',
                               'Item_id_d',
                               'Process',
                               'Copied',
                               'Error'])

    df['Barcode'] = barcodes
    df['Copied'] = False

# Display introduction
print(f'''

#####################
# Transfer IZ to IZ #
#####################

Configuration
=============

Environment: {env}
Source IZ: {iz_s}
Destination IZ: {iz_d}
Nb items: {len(df)}

Start job
=========

''')


######################
# Start copy of data #
######################

for i, barcode in enumerate(df['Barcode'].values):

    time.sleep(0.3)

    logging.info(f'{i+1} / {len(df["Barcode"].values)}: Handling {barcode}')

    # Skip row if already processed
    if len(df.loc[(df['Barcode'] == barcode) & (df['Copied'])]) > 0:
        continue

    # Fetch item data
    item_s = Item(barcode=barcode, zone=iz_s, env=env)
    item_id_s = item_s.get_item_id()
    holding_id_s = item_s.get_holding_id()
    mms_id_s = item_s.get_mms_id()
    nz_mms_id = item_s.get_nz_mms_id()

    # Save item data
    item_s.save()

    # Skip the row if error on the item
    if item_s.error is True:
        continue

    # Bib record
    # ----------

    # Check if copy bib record is required
    if len(df.loc[df['MMS_id_s'] == mms_id_s]) > 0:
        mms_id_d = df.loc[df['MMS_id_s'] == mms_id_s, 'MMS_id_d'].values[0]
        bib_d = IzBib(nz_mms_id, zone=iz_d, env=env, from_nz_mms_id=True)
    else:
        bib_d = IzBib(nz_mms_id, zone=iz_d, env=env, from_nz_mms_id=True, copy_nz_rec=True)
        mms_id_d = bib_d.get_mms_id()

    if bib_d.error is True:
        continue

    df.loc[df.Barcode == barcode, 'MMS_id_s'] = mms_id_s
    df.loc[df.MMS_id_s == mms_id_s, 'MMS_id_d'] = mms_id_d
    df.loc[df.Barcode == barcode, 'NZ_mms_id'] = nz_mms_id
    df.to_csv(process_file_path, index=False)

    # Holding record
    # --------------

    # Check if copy holding is required
    if len(df.loc[df['Holding_id_s'] == holding_id_s]) > 0:

        # Holding already created
        holding_id_d = df.loc[df['Holding_id_s'] == holding_id_s, 'Holding_id_d'].values[0]
    else:
        item_s.holding.save()

        # Get location according to the provided table
        loc_temp = locations_table.loc[(locations_table['Source library code'] == item_s.holding.library) &
                                       (locations_table['Source location code'] == item_s.holding.location),
                                       ['Destination library code', 'Destination location code']]

        if len(loc_temp) == 0:
            # Check if default location is available
            loc_temp = locations_table.loc[(locations_table['Source library code'] == '*DEFAULT*') &
                                           (locations_table['Source location code'] == '*DEFAULT*') &
                                           (~pd.isnull(locations_table['Destination library code'])) &
                                           (~pd.isnull(locations_table['Destination location code'])),
                                           ['Destination library code', 'Destination location code']]

        if len(loc_temp) == 0:
            # No corresponding location found => error
            logging.error(f'Location {item_s.holding.library}/{item_s.holding.location} not in locations table')
            continue

        # Get library and location destination
        library_d = loc_temp['Destination library code'].values[0]
        location_d = loc_temp['Destination location code'].values[0]

        # Load data of the source holding
        holding_temp = deepcopy(item_s.holding)

        # Change location and library
        holding_temp.location = location_d
        holding_temp.library = library_d

        # Get callnumber of the source holding
        callnumber_s = holding_temp.callnumber

        holding_d = None

        # Check if exists an destination holding with the same callnumber
        for holding in bib_d.get_holdings():
            callnumber_d = holding.callnumber

            if callnumber_d is not None and callnumber_s.strip() is not None \
                    and callnumber_d.strip() == callnumber_s.strip():
                logging.warning(f'{repr(item_s)}: holding found with same callnumber "{callnumber_s}"')
                holding_d = holding
                break

        if holding_d is None:
            # No holding found => need to be created
            holding_d = Holding(mms_id=mms_id_d, zone=iz_d, env=env, data=holding_temp.data, create_holding=True)

        holding_d.save()
        if holding_d.error is True:
            if 'Holding for this title at this location already exists' in holding_d.error_msg:
                error_label = 'similar_holding_existing'
            else:
                error_label = 'unknown_holding_error'

            df.loc[df.Barcode == barcode, 'Error'] = error_label
            continue
        holding_id_d = holding_d.get_holding_id()

    df.loc[df.Barcode == barcode, 'Holding_id_s'] = holding_id_s
    df.loc[df.Holding_id_s == holding_id_s, 'Holding_id_d'] = holding_id_d
    df.to_csv(process_file_path, index=False)

    # Create item
    # -----------
    loc_temp = locations_table.loc[(locations_table['Source library code'] == item_s.library) &
                                   (locations_table['Source location code'] == item_s.location),
                                   ['Destination library code', 'Destination location code']]

    if len(loc_temp) == 0:
        # Check if default location is available
        loc_temp = locations_table.loc[(locations_table['Source library code'] == '*DEFAULT*') &
                                       (locations_table['Source location code'] == '*DEFAULT*') &
                                       (~pd.isnull(locations_table['Destination library code'])) &
                                       (~pd.isnull(locations_table['Destination location code'])),
                                       ['Destination library code', 'Destination location code']]

    if len(loc_temp) == 0:
        # No corresponding location found => error
        logging.error(f'Location {item_s.library}/{item_s.location} not in locations table')
        continue

    # Get the new location and library of the item
    library_d = loc_temp['Destination library code'].values[0]
    location_d = loc_temp['Destination location code'].values[0]

    # Get the item policy
    policy_s = item_s.data.find('.//policy').text

    policy_temp = item_policies_table.loc[item_policies_table['Source item policy code'] == policy_s]

    # Check if default policy is available
    if len(policy_temp) == 0:
        policy_temp = item_policies_table.loc[item_policies_table['Source item policy code'] == '*DEFAULT*']

    if len(policy_temp) == 0:
        # No corresponding item policy found => error
        logging.error(f'Item policy {policy_s} not in item policies table')
        continue

    policy_d = policy_temp['Destination item policy code'].values[0]

    # Prepare the new item with a copy of the source item
    item_temp = deepcopy(item_s)
    item_temp.location = location_d
    item_temp.library = library_d
    item_temp.data.find('.//policy').text = policy_d

    # Clean blocking fields
    if FORCE_COPY is True:
        for field_name in ['provenance', 'temp_location', 'temp_library', 'in_temp_location', 'pattern_type',
                           'statistics_note_1', 'statistics_note_2', 'statistics_note_3', 'po_line']:
            fields = item_temp.data.findall(f'.//{field_name}')
            for field in fields:
                if field.text is not None or (field.text != 'false' and field_name == 'in_temp_location'):
                    logging.warning(f'{repr(item_temp)}: remove field "{field_name}", content: "{field.text}"')
                    field.getparent().remove(field)

    item_d = Item(mms_id_d, holding_id_d, zone=iz_d, env=env, data=item_temp.data, create_item=True)

    # Error handling => skip remaining process
    if item_d.error is True:
        if f'barcode {item_temp.barcode} already exists' in item_d.error_msg:
            # Get item by barcode
            item_d = Item(barcode=item_temp.barcode, zone=iz_d, env=env)
            error_label = 'already_exist'
        elif 'Given field provenance has invalid value' in item_d.error_msg:
            error_label = 'provenance_field'
        elif 'Request failed: Invalid temp_library code' in item_d.error_msg:
            error_label = 'temp_library'
        elif 'pattern_type is invalid' in item_d.error_msg:
            error_label = 'pattern_type'
        elif 'No response from Alma' in item_d.error_msg:
            item_d = Item(barcode=item_temp.barcode, zone=iz_d, env=env)
            if item_d.error is True:
                error_label = 'error_503_failed_to_create'
            else:
                error_label = 'error_503_success_to_create'
        else:
            error_label = 'unknown_item_error'
        df.loc[df.Barcode == barcode, 'Error'] = error_label

        # Skip remaining process
        if error_label not in ['already_exist', 'error_503_success_to_create']:
            continue

    # item_d.save()

    df.loc[df.Barcode == barcode, 'Item_id_s'] = item_s.get_item_id()
    df.loc[df.Barcode == barcode, 'Item_id_d'] = item_d.get_item_id()

    # Change barcode of source item
    if item_s.barcode.startswith('OLD_'):
        # Skip this step if barcode already updated
        logging.warning(f'{repr(item_d)}: barcode already updated "{item_d.barcode}"')
        continue

    item_s.barcode = 'OLD_' + item_s.barcode

    # Clean source item
    if FORCE_UPDATE is True:
        for field_name in ['pattern_type']:
            fields = item_s.data.findall(f'.//{field_name}')
            for field in fields:
                if field.text is not None:
                    logging.warning(f'{repr(item_temp)}: remove field "{field_name}", content: "{field.text}"')
                    field.getparent().remove(field)

    item_s.update()

    df.loc[df.Barcode == barcode, 'Copied'] = True
    df.to_csv(process_file_path, index=False)
