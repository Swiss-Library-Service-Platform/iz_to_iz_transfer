##############################
# Transfer IZ to IZ holdings #
##############################

# This script transfers item from IZ source to IZ destination.
# The information about the transfer should be given in an Excel file
# This file should be compliant with a given format

# To start the script:
# python transfer_iz_to_iz_holdings.py <dataForm.xlsx>

# Import libraries
from almapiwrapper.inventory import IzBib, Holding, Item
from almapiwrapper.configlog import config_log
import pandas as pd
from copy import deepcopy
import sys
import os
import logging
import openpyxl

# Config logs
config_log(sys.argv[1].replace('\\', '/').split('/')[-1].split('.')[0])

if len(sys.argv) != 2:
    # Bad argument, program stop
    logging.critical('Argument missing or not correct')
    exit()

# File path to the source data Excel file
src_data_file_path = sys.argv[1]

# File path to the backup of the processed records
process_file_path = 'data/' + src_data_file_path.replace('\\', '/').split('/')[-1].split('.')[0] + '_holdings_processing.csv'

# Get configuration
wb = openpyxl.load_workbook(src_data_file_path)
wb.active = wb['General']
sheet = wb.active
iz_s = sheet.cell(row=3, column=2).value
iz_d = sheet.cell(row=4, column=2).value
env = {'Production': 'P',
       'Sandbox': 'S'}.get(sheet.cell(row=5, column=2).value, 'P')
FORCE_COPY = {'Yes': True, 'No': False}.get(sheet.cell(row=7, column=2).value, False)
FORCE_UPDATE = {'Yes': True, 'No': False}.get(sheet.cell(row=8, column=2).value, False)

# Load holding_ids and MMS Ids of source records
src_data = pd.read_excel(src_data_file_path, sheet_name='Holdings', dtype=str).dropna()

holding_ids = src_data['Holding_id'].str.strip("'")
mms_ids = src_data['IZ_MMS_id'].str.strip("'")
logging.info(f'{len(holding_ids)} holding ids loaded from "{src_data_file_path}" file.')

# Load locations
locations_table = pd.read_excel(src_data_file_path, sheet_name='Locations_mapping', dtype=str)

# Check if processing file exists
if os.path.exists(process_file_path) is True:
    df = pd.read_csv(process_file_path, dtype=str)
    df = df.replace('False', False)
    df = df.replace('True', True)
    df = df.replace('NaN', None)

else:
    df = pd.DataFrame(columns=['NZ_mms_id',
                               'MMS_id_s',
                               'Holding_id_s',
                               'MMS_id_d',
                               'Holding_id_d',
                               'Process',
                               'Copied',
                               'Error'])
    df['Holding_id_s'] = holding_ids
    df['MMS_id_s'] = mms_ids
    df['Copied'] = False

# Display introduction
print(f'''

##############################
# Transfer IZ to IZ holdings #
##############################

Configuration
=============

Environment: {env}
Source IZ: {iz_s}
Destination IZ: {iz_d}
Nb holdings: {len(df)}

Start job
=========

''')


######################
# Start copy of data #
######################

for i, holding_id_s in enumerate(df['Holding_id_s'].values):

    logging.info(f'{i+1} / {len(df["Holding_id_s"].values)}: Handling {holding_id_s}')

    # Skip row if already processed
    if len(df.loc[(df['Holding_id_s'] == holding_id_s) & (df['Copied'])]) > 0:
        continue

    # Get mms_id of source record
    mms_id_s = df.loc[df['Holding_id_s'] == holding_id_s, 'MMS_id_s'].values[0]

    # Fetch holding data
    holding_s = Holding(mms_id_s, holding_id_s, zone=iz_s, env=env)

    # Save item data
    holding_s.save()

    # Skip the row if error on the item
    if holding_s.error is True or holding_s.bib.get_mms_id() != mms_id_s:
        if holding_s.error is True:
            error_label = 'Error by fetching source holding'
        else:
            error_label = f'mms_id ({holding_s.bib.get_mms_id()}) of record linked to holding not same as provided mms_id'
            holding_s.error = True
        df.loc[df['Holding_id_s'] == holding_id_s, 'Error'] = error_label
        df.to_csv(process_file_path, index=False)
        continue

    # Get local bib record and NZ mms_id
    nz_mms_id = holding_s.bib.get_nz_mms_id()


    # Bib record
    # ----------

    # Check if copy bib record is required
    if len(df.loc[df['MMS_id_s'] == mms_id_s, 'MMS_id_d'].dropna().values) > 0:
        mms_id_d = df.loc[df['MMS_id_s'] == mms_id_s, 'MMS_id_d'].values[0]
        bib_d = IzBib(nz_mms_id, zone=iz_d, env=env, from_nz_mms_id=True)
    else:
        bib_d = IzBib(nz_mms_id, zone=iz_d, env=env, from_nz_mms_id=True, copy_nz_rec=True)
        mms_id_d = bib_d.get_mms_id()

    if bib_d.error is True:
        error_label = 'Unable to get a destination bib record'
        df.loc[df.Holding_id_s == holding_id_s, 'Error'] = error_label
        continue

    df.loc[df.Holding_id_s == holding_id_s, 'MMS_id_d'] = mms_id_d
    df.loc[df.Holding_id_s == holding_id_s, 'NZ_mms_id'] = nz_mms_id
    df.to_csv(process_file_path, index=False)

    # Holding record
    # --------------

    # Check if copy holding is required
    if len(df.loc[(df['Holding_id_s'] == holding_id_s) & (pd.notna(df['Holding_id_d']))]) > 0:

        # Holding already created
        holding_id_d = df.loc[df['Holding_id_s'] == holding_id_s, 'Holding_id_d'].values[0]
    else:

        # Get location according to the provided table
        loc_temp = locations_table.loc[(locations_table['Source library code'] == holding_s.library) &
                                       (locations_table['Source location code'] == holding_s.location),
                                       ['Destination library code', 'Destination location code']]

        if len(loc_temp) == 0:
            # Check if default location is available
            loc_temp = locations_table.loc[(locations_table['Source library code'] == '*DEFAULT*') &
                                           (locations_table['Source location code'] == '*DEFAULT*') &
                                           (~pd.isnull(locations_table['Destination library code'])) &
                                           (~pd.isnull(locations_table['Destination location code'])),
                                           ['Destination library code', 'Destination location code']]

        if len(loc_temp) == 0:
            # No corresponding location found => error
            logging.error(f'Location {holding_s.library}/{holding_s.location} not in locations table')
            error_label = 'Location not existing in location table'
            df.loc[df['Holding_id_s'] == holding_id_s, 'Error'] = error_label
            continue

        # Get library and location destination
        library_d = loc_temp['Destination library code'].values[0]
        location_d = loc_temp['Destination location code'].values[0]

        # Load data of the source holding
        holding_temp = deepcopy(holding_s)

        # Change location and library
        holding_temp.location = location_d
        holding_temp.library = library_d

        # Get callnumber of the source holding
        callnumber_s = holding_temp.callnumber

        # Suppress empty chars from call numbers
        if callnumber_s is not None:
            callnumber_s.strip()

        holding_d = None

        # Check if exists a destination holding with the same callnumber
        for holding in bib_d.get_holdings():
            callnumber_d = holding.callnumber

            # Suppress empty chars from call numbers
            if callnumber_d is not None:
                callnumber_d = callnumber_d.strip()

            if callnumber_d == callnumber_s:
                logging.error(f'{repr(holding_s)}: holding found with same callnumber "{callnumber_s}"')
                holding_d = holding
                holding_d.error = True
                holding_d.error_msg = f'Holding for this title at this location already exists "{callnumber_s}"'

        if holding_d is None:
            # No holding found => need to be created
            holding_d = Holding(mms_id=mms_id_d, zone=iz_d, env=env, data=holding_temp.data, create_holding=True)

        holding_d.save()
        if holding_d.error is True:
            if 'Holding for this title at this location already exists' in holding_d.error_msg:
                error_label = 'similar_holding_existing'
                df.loc[df['Holding_id_s'] == holding_id_s, 'Error'] = error_label
                holding_id_d = holding_d.holding_id
            else:
                error_label = 'unknown_holding_error'
                df.loc[df['Holding_id_s'] == holding_id_s, 'Error'] = error_label
                continue
        else:
            holding_id_d = holding_d.get_holding_id()

    df.loc[df.Holding_id_s == holding_id_s, 'Holding_id_d'] = holding_id_d
    df.to_csv(process_file_path, index=False)

    df.loc[df.Holding_id_s == holding_id_s, 'Copied'] = True
    df.to_csv(process_file_path, index=False)

# Make a report with the errors
df.loc[(~df['Copied']) | (df['Error'])].to_csv(process_file_path.replace('_processing.csv', '_not_copied.csv'),
                                               index=False)
