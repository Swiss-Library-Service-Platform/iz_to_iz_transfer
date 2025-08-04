import openpyxl
import pandas as pd
import os

from typing import Tuple, Optional

# Global variable to store process configuration, it is initialized in set_config
_config_cache = {}


def get_raw_filename(filepath: str) -> str:
    """
    Extracts the base filename without its extension from a given file path.

    Parameters
    ----------
    filepath : str
        Full or relative path to the file.

    Returns
    -------
    str
        Filename without extension.
    """
    return os.path.splitext(os.path.basename(filepath))[0]


def get_form_version(excel_filepath: str) -> str:
    """
    Reads the version of the Excel form from the 'General' tab of the provided Excel file.

    Parameters
    ----------
    excel_filepath : str
        Path to the Excel file containing configuration data.

    Returns
    -------
    str
        Version of the Excel form.
    """
    wb = openpyxl.load_workbook(excel_filepath)
    wb.active = wb['General']
    sheet = wb.active

    version = sheet.cell(row=5, column=2).value

    return version


def set_config(excel_filepath: str):
    """
    Reads configuration from an Excel file and returns it as a dictionary.

    Parameters
    ----------
    excel_filepath : str
        Path to the Excel file containing configuration data.

    Returns
    -------
    dict
        Dictionary containing configuration data.
    """
    global _config_cache

    wb = openpyxl.load_workbook(excel_filepath)

    # Get General tab information
    wb.active = wb['General']
    sheet = wb.active

    config = {
        'iz_s': sheet.cell(row=6, column=2).value,
        'iz_d': sheet.cell(row=7, column=2).value,
        'lib_s': sheet.cell(row=8, column=2).value,
        'lib_d': sheet.cell(row=9, column=2).value,
        'env': {'Production': 'P', 'Sandbox': 'S'}.get(sheet.cell(row=10, column=2).value, 'P'),
        'acq_department': sheet.cell(row=11, column=2).value,
        'make_reception': True if sheet.cell(row=12, column=2).value == 'Yes' else False,
        'interested_users': [],
        'items_fields': {'src': {'to_delete': [], 'to_delete_if_error': []},
        'dest': {'to_delete': [], 'to_delete_if_error': []}},
        'polines_fields': {'to_delete': [], 'to_delete_if_error': []}
    }

    # Read items fields to delete from the Excel sheet
    for i in range(15, 22):
        key = sheet.cell(row=i, column=1).value
        value_src = sheet.cell(row=i, column=2).value
        value_dest = sheet.cell(row=i, column=3).value

        if value_src == 'Always delete':
            config['items_fields']['src']['to_delete'] += key.split(', ')
        elif value_src == 'Delete if error':
            config['items_fields']['src']['to_delete_if_error'] += key.split(', ')

        if value_dest == 'Always delete':
            config['items_fields']['dest']['to_delete'] += key.split(', ')
        elif value_dest == 'Delete if error':
            config['items_fields']['dest']['to_delete_if_error'] += key.split(', ')

    # Read polines fields to delete from the Excel sheet
    for i in range(24, 25):
        key = sheet.cell(row=i, column=1).value
        value = sheet.cell(row=i, column=3).value

        if value == 'Always delete':
            config['polines_fields']['to_delete'] += key.split(', ')
        elif value == 'Delete if error':
            config['polines_fields']['to_delete_if_error'] += key.split(', ')


    # Get Locations_mapping tab information
    config['locations_mapping'] = pd.read_excel(excel_filepath, sheet_name='Locations_mapping', dtype=str)

    # Get item policies mapping tab information
    # config['item_policies_mapping'] = pd.read_excel(excel_filepath, sheet_name='Item_policies_mapping', dtype=str)

    # Get vendors mapping tab information
    config['vendors_mapping'] = pd.read_excel(excel_filepath, sheet_name='Vendors_mapping', dtype=str)

    # Get funding sources mapping tab information
    config['Funds_mapping'] = pd.read_excel(excel_filepath, sheet_name='Funds_mapping', dtype=str)

    _config_cache = config

def get_config() -> dict:
    """
    Returns the cached configuration dictionary.

    Returns
    -------
    dict
        Cached configuration dictionary.
    """
    global _config_cache
    return _config_cache

def get_data(excel_filepath: str, sheet_name) -> pd.DataFrame:
    """
    Reads the specified tab from an Excel file and returns a DataFrame containing the information from the order lines.

    Parameters
    ----------
    excel_filepath : str
        Path to the Excel file containing the data.
    sheet_name : str
        Name of the tab to read.

    Returns
    -------
    pd.DataFrame
        DataFrame with order line numbers, MMS ID, holding ID, etc.
    """
    return pd.read_excel(excel_filepath, sheet_name=sheet_name, dtype=str)

def get_corresponding_location(library_s: str, location_s: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Returns the corresponding library and location for a given library and location.

    Parameters
    ----------
    library_s : str
        The library name.
    location_s : str
        The location name.

    Returns
    -------
    Tuple[str, str] or None, None
        A tuple containing the corresponding library and location.
        Returns Tuple of None if no corresponding location is found.
    """
    locations_table = get_config()['locations_mapping']

    loc_temp = locations_table.loc[(locations_table['Source library code'] == library_s) &
                                   (locations_table['Source location code'] == location_s),
                                   ['Destination library code', 'Destination location code']]

    if len(loc_temp) == 0:
        # Check if default location is available
        loc_temp = locations_table.loc[(locations_table['Source library code'] == '*DEFAULT*') &
                                       (locations_table['Source location code'] == '*DEFAULT*') &
                                       (~pd.isnull(locations_table['Destination library code'])) &
                                       (~pd.isnull(locations_table['Destination location code'])),
                                       ['Destination library code', 'Destination location code']]

    if len(loc_temp) == 0:
        # No corresponding location found => error
        return None, None

    # Get the new location and library of the item
    library_d = loc_temp['Destination library code'].values[0]
    location_d = loc_temp['Destination location code'].values[0]

    return library_d, location_d

def get_corresponding_library(library_s: str) -> Optional[str]:
    """
    Returns the corresponding library for a given library.

    Parameters
    ----------
    library_s : str
        The library name.

    Returns
    -------
    str or None
        The corresponding library name, or None if no corresponding library is found.
    """


    locations_table = get_config()['locations_mapping']

    loc_temp = locations_table.loc[locations_table['Source library code'] == library_s,
    ['Destination library code']]

    if len(loc_temp) == 0:
        # Check if default location is available
        loc_temp = locations_table.loc[((locations_table['Source library code'] == '*DEFAULT*') &
                                       (~pd.isnull(locations_table['Destination library code']))),
        ['Destination library code']]

    if len(loc_temp) == 0:
        # No corresponding location found => error
        return None

    # Get the new location and library of the item
    library_d = loc_temp['Destination library code'].values[0]

    return library_d

# def get_corresponding_item_policy(item_policy_s: str) -> Optional[str]:
#     """
#     Returns the corresponding item policy for a given item policy.
#     Parameters
#     ----------
#     item_policy_s : str
#         The item policy name.
#
#     Returns
#     -------
#     str or None
#         The corresponding item policy name, or None if no corresponding item policy is found.
#     """
#
#     item_policies_table = get_config()['item_policies_mapping']
#
#     item_policy_temp = item_policies_table.loc[item_policies_table['Source item policy'] == item_policy_s,
#                                                'Destination item policy']
#
#     if len(item_policy_temp) == 0:
#         # Check if default item policy is available
#         item_policy_temp = item_policies_table.loc[(item_policies_table['Source item policy'] == '*DEFAULT*') &
#                                                    (~pd.isnull(item_policies_table['Destination item policy'])),
#                                                    'Destination item policy']
#
#     if len(item_policy_temp) == 0:
#         # No corresponding item policy found => error
#         return None
#
#     # Get the new item policy
#     item_policy_d = item_policy_temp.values[0]
#
#     return item_policy_d


def get_corresponding_vendor(vendor_s: str, vendor_account_s: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Returns the corresponding vendor and vendor account for a given vendor name and vendor account.

    Parameters
    ----------
    vendor_s : str
        The vendor code of the source library.
    vendor_account_s : str
        The vendor account code of the source library.

    Returns
    -------
    Tuple[str, str] or None
        A tuple containing the corresponding vendor and vendor account.
        Returns None if no corresponding vendor is found.
    """
    vendors_table = get_config()['vendors_mapping']

    vendor_temp = vendors_table.loc[(vendors_table['Source vendor code'] == vendor_s) &
                                    (vendors_table['Source vendor account'] == vendor_account_s),
                                    ['Destination vendor code', 'Destination vendor account']]

    if len(vendor_temp) == 0:
        # Check if default vendor is available
        vendor_temp = vendors_table.loc[(vendors_table['Source vendor code'] == vendor_s) &
                                        (vendors_table['Source vendor account'] == '*DEFAULT*') &
                                        (~pd.isnull(vendors_table['Destination vendor code'])) &
                                        (~pd.isnull(vendors_table['Destination vendor account'])),
                                        ['Destination vendor code', 'Destination vendor account']]

    if len(vendor_temp) == 0:
        # Check if default vendor is available
        vendor_temp = vendors_table.loc[(vendors_table['Source vendor code'] == '*DEFAULT*') &
                                        (vendors_table['Source vendor account'] == '*DEFAULT*') &
                                        (~pd.isnull(vendors_table['Destination vendor code'])) &
                                        (~pd.isnull(vendors_table['Destination vendor account'])),
                                        ['Destination vendor code', 'Destination vendor account']]

    if len(vendor_temp) == 0:
        # No corresponding vendor found => error
        return None, None

    # Get the new vendor and vendor account
    vendor_d = vendor_temp['Destination vendor code'].values[0]
    vendor_account_d = vendor_temp['Destination vendor account'].values[0]

    return vendor_d, vendor_account_d


def get_corresponding_fund(fund_code_s: str) -> Optional[str]:
    """
    Returns the corresponding fund code for a given fund code.

    Parameters
    ----------
    fund_code_s : str
        The fund code of the source library.

    Returns
    -------
    str or None
        The corresponding fund code, or None if no corresponding fund code is found.
    """
    funds_table = get_config()['Funds_mapping']

    fund_temp = funds_table.loc[funds_table['Source fund code'] == fund_code_s, 'Destination fund code']

    if len(fund_temp) == 0:
        # Check if default fund is available
        fund_temp = funds_table.loc[(funds_table['Source fund code'] == '*DEFAULT*') &
                                    (~pd.isnull(funds_table['Destination fund code'])),
                                    'Destination fund code']

    if len(fund_temp) == 0:
        # No corresponding fund found => error
        return None

    # Get the new fund code
    fund_code_d = fund_temp.values[0]

    return fund_code_d